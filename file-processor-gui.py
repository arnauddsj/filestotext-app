import sys
import os
import csv
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLineEdit, QLabel, QFileDialog, QTextEdit, QMessageBox
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont
import fnmatch
import requests
import openpyxl
from docx import Document
import PyPDF2

VERSION = "1.2.0"
VERSION_URL = "https://example.com/version.txt"

class VersionCheckerThread(QThread):
    version_checked = pyqtSignal(str)

    def run(self):
        try:
            response = requests.get(VERSION_URL, timeout=5)
            if response.status_code == 200:
                latest_version = response.text.strip()
                self.version_checked.emit(latest_version)
        except:
            pass  # Silently fail if unable to check version

class ClearableLineEdit(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        self.lineEdit = QLineEdit(self)
        self.clearButton = QPushButton("×", self)
        self.clearButton.setFixedSize(20, 20)
        self.clearButton.clicked.connect(self.lineEdit.clear)
        layout.addWidget(self.lineEdit)
        layout.addWidget(self.clearButton)

    def text(self):
        return self.lineEdit.text()

    def setText(self, text):
        self.lineEdit.setText(text)

class FileProcessorApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.check_version()

    def initUI(self):
        layout = QVBoxLayout()

        # Version label
        self.version_label = QLabel(f"Version: {VERSION}")
        self.version_label.setAlignment(Qt.AlignRight)
        layout.addWidget(self.version_label)

        # Input folder selection
        input_layout = QHBoxLayout()
        self.input_edit = QLineEdit()
        input_button = QPushButton("Select Input Folder")
        input_button.clicked.connect(self.select_input_folder)
        input_layout.addWidget(QLabel("Input Folder:"))
        input_layout.addWidget(self.input_edit)
        input_layout.addWidget(input_button)
        layout.addLayout(input_layout)

        # Output folder selection
        output_layout = QHBoxLayout()
        self.output_edit = QLineEdit()
        output_button = QPushButton("Select Output Folder")
        output_button.clicked.connect(self.select_output_folder)
        output_layout.addWidget(QLabel("Output Folder:"))
        output_layout.addWidget(self.output_edit)
        output_layout.addWidget(output_button)
        layout.addLayout(output_layout)

        # Ignore patterns
        layout.addWidget(QLabel("Ignore Files (comma-separated):"))
        self.ignore_files_edit = ClearableLineEdit()
        self.ignore_files_edit.setText("*.log,*.tmp,*.cache,.DS_Store,*.py,.env,package-lock.json,*.svg,*.ico")
        layout.addWidget(self.ignore_files_edit)

        layout.addWidget(QLabel("Ignore Directories (comma-separated):"))
        self.ignore_dirs_edit = ClearableLineEdit()
        self.ignore_dirs_edit.setText("node_modules,build,dist,migrations,venv,.git")
        layout.addWidget(self.ignore_dirs_edit)

        # Process button
        self.process_button = QPushButton("Process Files")
        self.process_button.clicked.connect(self.process_files)
        layout.addWidget(self.process_button)

        # Log output
        self.log_output = QTextEdit()
        self.log_output.setReadOnly(True)
        layout.addWidget(self.log_output)

        self.setLayout(layout)
        self.setWindowTitle('File Processor')
        self.setGeometry(300, 300, 500, 400)

    def check_version(self):
        self.version_thread = VersionCheckerThread()
        self.version_thread.version_checked.connect(self.on_version_checked)
        self.version_thread.start()

    def on_version_checked(self, latest_version):
        if latest_version > VERSION:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText("A new version is available!")
            msg.setInformativeText(f"Current version: {VERSION}\nLatest version: {latest_version}")
            msg.setWindowTitle("Update Available")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()

    def select_input_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Input Folder")
        if folder:
            self.input_edit.setText(folder)

    def select_output_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if folder:
            self.output_edit.setText(folder)

    def process_files(self):
        input_folder = self.input_edit.text()
        output_folder = self.output_edit.text() or input_folder
        ignore_files = [pattern.strip() for pattern in self.ignore_files_edit.text().split(',') if pattern.strip()]
        ignore_dirs = [dir.strip() for dir in self.ignore_dirs_edit.text().split(',') if dir.strip()]

        if not input_folder:
            self.log_output.append("Please select an input folder.")
            return

        self.log_output.clear()
        self.log_output.append(f"Processing files in: {input_folder}")
        self.log_output.append(f"Output folder: {output_folder}")

        try:
            file_paths = self.get_file_paths(input_folder, ignore_files, ignore_dirs)
            if not file_paths:
                self.log_output.append("No files found to process.")
                return

            output_file_path = os.path.join(output_folder, f"{os.path.basename(input_folder)}_files.txt")
            processed_count, error_count = self.write_to_txt(output_file_path, file_paths)
            
            self.log_output.append(f"Processed {processed_count} files.")
            self.log_output.append(f"Encountered {error_count} errors.")
            self.log_output.append(f"Output written to: {output_file_path}")
        except Exception as e:
            self.log_output.append(f"An error occurred: {str(e)}")

    def get_file_paths(self, start_path, ignore_files, ignore_dirs):
        file_paths = []
        for root, dirs, files in os.walk(start_path, topdown=True):
            dirs[:] = [d for d in dirs if d not in ignore_dirs]
            for file in files:
                if not any(fnmatch.fnmatch(file, pattern) for pattern in ignore_files):
                    full_path = os.path.join(root, file)
                    file_paths.append(full_path)
        return file_paths

    def write_to_txt(self, output_path, file_paths):
        processed_count = 0
        error_count = 0
        with open(output_path, 'w', encoding='utf-8') as output_file:
            for path in file_paths:
                try:
                    content = self.read_file_content(path)
                    output_file.write(f"PATH: {path}\nCONTENT:\n{content}\n\n")
                    processed_count += 1
                except Exception as e:
                    error_count += 1
                    self.log_output.append(f"Error processing {path}: {e}")
                    output_file.write(f"PATH: {path}\nCONTENT: [Error reading file]\n\n")
        return processed_count, error_count

    def read_file_content(self, file_path):
        _, ext = os.path.splitext(file_path.lower())
        
        if ext == '.csv':
            return self.read_csv(file_path)
        elif ext in ['.xlsx', '.xls']:
            return self.read_excel(file_path)
        elif ext == '.docx':
            return self.read_docx(file_path)
        elif ext == '.pdf':
            return self.read_pdf(file_path)
        elif self.is_binary(file_path):
            return "[Binary file, content not included]"
        else:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()

    def read_csv(self, file_path):
        with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            return '\n'.join([','.join(row) for row in reader])

    def read_excel(self, file_path):
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        content = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            content.append(f"Sheet: {sheet}")
            for row in ws.iter_rows(values_only=True):
                content.append(','.join(str(cell) for cell in row if cell is not None))
        return '\n'.join(content)

    def read_docx(self, file_path):
        doc = Document(file_path)
        return '\n'.join([paragraph.text for paragraph in doc.paragraphs])

    def read_pdf(self, file_path):
        content = []
        with open(file_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            for page in reader.pages:
                content.append(page.extract_text())
        return '\n'.join(content)

    def is_binary(self, file_path, chunk_size=1024):
        try:
            with open(file_path, 'rb') as file:
                return b'\0' in file.read(chunk_size)
        except Exception:
            return True  # If we can't read the file, assume it's binary for safety

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = FileProcessorApp()
    ex.show()
    sys.exit(app.exec_())